# import json
# import logging
# from django.http import HttpResponse, JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# from django.db import connection, IntegrityError
# from openpyxl import load_workbook
# from datetime import datetime

# logger = logging.getLogger(__name__)

# @csrf_exempt
# def add_exam(request):
#     if request.method == 'POST':
#         try:
#             if 'application/json' in request.content_type:
#                 # Handle JSON data
#                 json_data = json.loads(request.body)
#                 title = json_data.get('title')
#                 start_date = json_data.get('start_date')
#                 description = json_data.get('description')
#                 # Parse time and convert to 'HH:MM:SS' format
#                 time_str = json_data.get('time')
#                 time = datetime.strptime(time_str, '%I:%M %p').strftime('%H:%M:%S')
#                 duration = json_data.get('duration')
#                 marks = json_data.get('marks')
#                 rules = json_data.get('rules')
#                 creator = json_data.get('creator')
#                 select_students = json_data.get('select_students', [])  # If not provided, default to empty list
#                 excel_file = None  # Initialize excel_file to None
#             else:
#                 # Handle FormData
#                 title = request.POST.get('title')
#                 start_date = request.POST.get('start_date')
#                 description = request.POST.get('description')
#                 # Parse time and convert to 'HH:MM:SS' format
#                 time_str = request.POST.get('time')
#                 time = datetime.strptime(time_str, '%I:%M %p').strftime('%H:%M:%S')
#                 duration = request.POST.get('duration')
#                 marks = request.POST.get('marks')
#                 rules = request.POST.get('rules')
#                 creator = request.POST.get('creator')
#                 select_students = request.POST.get('select_students')
#                 select_students = [int(user_id) for user_id in select_students.split(',')] if select_students else []

#                 # Fetching uploaded file
#                 excel_file = request.FILES.get('file')

#             # Input validation
#             if not all([title, start_date, time, duration, marks, creator]):
#                 return HttpResponse("Missing required fields", status=400)

#             with connection.cursor() as cursor:
#                 cursor.execute("""
#                     SELECT user_id FROM user WHERE full_name = %s AND role = 2
#                 """, [creator])

#                 creator_data = cursor.fetchone()
#                 if creator_data:
#                     creator_user_id = creator_data[0]
#                 else:
#                     return HttpResponse("No user found for creator: {}".format(creator), status=404)

#                 cursor.execute("""
#                     INSERT INTO exams (title, start_date, description, time, duration, marks, rules, creator)
#                     VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
#                 """, [title, start_date, description, time, duration, marks, rules, creator_user_id])

#                 connection.commit()

#                 exam_id = cursor.lastrowid  # Use last inserted row id

#                 # Check if any students are selected
#                 if select_students:
#                     for select_student in select_students:
#                         cursor.execute("""
#                             INSERT INTO candidate_exam (exam_id, user_id, approval_status)
#                             VALUES (%s, %s, 'Y')
#                         """, [exam_id, select_student])
#                         logger.info("Candidate added for exam_id %s and user_id %s", exam_id, select_student)

#                 if excel_file:
#                     wb = load_workbook(excel_file)
#                     ws = wb.active

#                     for row in ws.iter_rows(min_row=2, values_only=True):
#                         full_name, email = row[:2]
#                         cursor.execute("""
#                             SELECT user_id FROM user WHERE email = %s
#                         """, [email])
#                         user_data = cursor.fetchone()
#                         if user_data:
#                             user_id = user_data[0]
#                             cursor.execute("""
#                                 INSERT INTO candidate_exam (exam_id, user_id, approval_status)
#                                 VALUES (%s, %s, 'Y')
#                             """, [exam_id, user_id])
#                             logger.info("Candidate added for exam_id %s and user_id %s", exam_id, user_id)
#                         else:
#                             logger.warning("User not found for email: %s", email)

#             response_data = {'status': 'Exam added successfully'}
#             return JsonResponse(response_data, status=200)

#         except IntegrityError as e:
#             logger.error("Integrity Error: %s", str(e))
#             return HttpResponse("Integrity Error: {}".format(e), status=400)
#         except Exception as e:
#             logger.error("An error occurred: %s", str(e))
#             logger.exception("Detailed exception:")
#             return HttpResponse("An error occurred: {}".format(e), status=500)


# Added Email send to selected students -- Sufiyan

import json
import logging
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import connection, IntegrityError
from openpyxl import load_workbook
from datetime import datetime

from django.core.mail import send_mail

logger = logging.getLogger(__name__)

# @csrf_exempt
# def add_exam(request):
#     if request.method == 'POST':
#         try:
#             response_data_email = None
#             if 'application/json' in request.content_type:
#                 # Handle JSON data
#                 json_data = json.loads(request.body)
#                 title = json_data.get('title')
#                 start_date = json_data.get('start_date')
#                 description = json_data.get('description')
#                 # Parse time and convert to 'HH:MM:SS' format
#                 time_str = json_data.get('time')
#                 time = datetime.strptime(time_str, '%I:%M %p').strftime('%H:%M:%S')
#                 duration = json_data.get('duration')
#                 marks = json_data.get('marks')
#                 rules = json_data.get('rules')
#                 creator = json_data.get('creator')
#                 select_students = json_data.get('select_students', [])  # If not provided, default to empty list
#                 excel_file = None  # Initialize excel_file to None
#             else:
#                 # Handle FormData
#                 title = request.POST.get('title')
#                 start_date = request.POST.get('start_date')
#                 description = request.POST.get('description')
#                 # Parse time and convert to 'HH:MM:SS' format
#                 time_str = request.POST.get('time')
#                 time = datetime.strptime(time_str, '%I:%M %p').strftime('%H:%M:%S')
#                 duration = request.POST.get('duration')
#                 marks = request.POST.get('marks')
#                 rules = request.POST.get('rules')
#                 creator = request.POST.get('creator')
#                 select_students = request.POST.get('select_students')
#                 select_students = [int(user_id) for user_id in select_students.split(',')] if select_students else []

#                 # Fetching uploaded file
#                 excel_file = request.FILES.get('file')

#             # Input validation
#             if not all([title, start_date, time, duration, marks, creator]):
#                 return HttpResponse("Missing required fields", status=400)

#             with connection.cursor() as cursor:
#                 cursor.execute("""
#                     SELECT user_id FROM user WHERE full_name = %s AND role = 2
#                 """, [creator])

#                 creator_data = cursor.fetchone()
#                 if creator_data:
#                     creator_user_id = creator_data[0]
#                 else:
#                     return HttpResponse("No user found for creator: {}".format(creator), status=404)

#                 cursor.execute("""
#                     INSERT INTO exams (title, start_date, description, time, duration, marks, rules, creator)
#                     VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
#                 """, [title, start_date, description, time, duration, marks, rules, creator_user_id])

#                 connection.commit()

#                 exam_id = cursor.lastrowid  # Use last inserted row id

#                 # Check if any students are selected
#                 if select_students:
#                     for select_student in select_students:
#                         cursor.execute("""
#                             INSERT INTO candidate_exam (exam_id, user_id, approval_status)
#                             VALUES (%s, %s, 'Y')
#                         """, [exam_id, select_student])
#                         logger.info("Candidate added for exam_id %s and user_id %s", exam_id, select_student)

#                 if excel_file:
#                     # Check if the uploaded file is in the expected Excel format
#                     if not excel_file.name.endswith('.xlsx'):
#                         return HttpResponse("Invalid file format. Please upload an Excel file (.xlsx)", status=400)

#                     # Attempt to load the workbook
#                     try:
#                         wb = load_workbook(excel_file)
#                     except Exception as e:
#                         logger.error("Failed to load Excel file: %s", str(e))
#                         return HttpResponse("Failed to load Excel file: {}".format(e), status=400)

#                     ws = wb.active

#                     # Continue processing the workbook data
#                     for row in ws.iter_rows(min_row=2, values_only=True):
#                         full_name, email = row[:2]
#                         cursor.execute("""
#                             SELECT user_id FROM user WHERE email = %s
#                         """, [email])
#                         user_data = cursor.fetchone()
#                         if user_data:
#                             user_id = user_data[0]
#                             cursor.execute("""
#                                 INSERT INTO candidate_exam (exam_id, user_id, approval_status)
#                                 VALUES (%s, %s, 'Y')
#                             """, [exam_id, user_id])
#                             logger.info("Candidate added for exam_id %s and user_id %s", exam_id, user_id)
#                         else:
#                             logger.warning("User not found for email: %s", email)
#             # sufiyan email send to students
#               # Get the list of selected student IDs

#             # select_students = data.get('select_students')
#             # select_students = select_students
#             student_ids = select_students
#             # if not student_ids:
#             #     return JsonResponse({'error': 'No students selected'}, status=400)
#             if student_ids:
#                 try:
#                     student_ids = [int(id) for id in student_ids]
#                 except (ValueError, TypeError):
#                     return JsonResponse({'error': 'Student IDs must be integers'}, status=400)
#                 # Convert the comma-separated string to a list of integers
#                 # student_ids = [int(id) for id in select_students.split(',')]

#                 # Fetch email addresses from the database
#                 with connection.cursor() as cursor:
#                     # Construct a parameterized query with the right number of placeholders
#                     placeholders = ', '.join(['%s'] * len(student_ids))
#                     query = f"SELECT user_id, email FROM user WHERE user_id IN ({placeholders})"

#                     # Execute the query with student IDs as parameters
#                     cursor.execute(query, student_ids)
#                     students = cursor.fetchall()

#                 # If no matching students found
#                 if not students:
#                     return JsonResponse({'error': 'No matching students found'}, status=404)

#                 # Prepare email details
#                 subject = f"Invitaion: {title}"
#                 message = f'''
#                 Dear Student,

#                 You are invited to participate in the following exam:

#                 Title: {title}
#                 Date: {start_date}
#                 Time: {time}
#                 Duration: {duration} minutes

#                 Description:
#                 {description}

#                 Please log in to your account for more details and to confirm your participation:
#                 https://aaas.tdtl.world/login/

#                 Best regards,
#                 Your Institution
#                 '''
#                 from_email = 'pr@thedatatechlabs.com'

#                 # Send email to each student
#                 for student_id, email in students:
#                     to_email = [email]
#                     try:
#                         send_mail(
#                             subject,
#                             message,
#                             from_email,
#                             to_email,
#                             fail_silently=False
#                         )
#                         print(f"Email sent successfully to student {student_id} at {email}")
#                     except Exception as e:
#                         print(f"Error sending email to student {student_id} at {email}: {e}")
#                 response_data_email = {
#                               'message': 'Emails sent successfully',
#                 'students_notified': [id for id, _ in students]}

#             # Return success response
#             # return JsonResponse({
#             #     'message': 'Emails sent successfully',
#             #     'students_notified': [id for id, _ in students]
#             # })
#             # sufiyan email send to students

#             # response_data = {'status': 'Exam added successfully',
#             #                  if response_data_email:
#             #                     response_data_email
#             #                   }
#             response_data = {'status': 'Exam added successfully'}

#             if response_data_email:
#                 response_data['email_info'] = response_data_email
#             # response_data = {'status': 'Exam added successfully',
#             #                   'message': 'Emails sent successfully',
#             #     'students_notified': [id for id, _ in students]}

#             return JsonResponse(response_data, status=200)

#         except IntegrityError as e:
#             logger.error("Integrity Error: %s", str(e))
#             return HttpResponse("Integrity Error: {}".format(e), status=400)
#         except Exception as e:
#             logger.error("An error occurred: %s", str(e))
#             logger.exception("Detailed exception:")
#             return HttpResponse("An error occurred: {}".format(e), status=500)


# **************************************************************************************************************************


@csrf_exempt
def add_exam(request):
    if request.method == 'POST':
        try:
            response_data_email = (
                None  # Initialize response_data_email to None at the start
            )

            if "application/json" in request.content_type:
                # Handle JSON data
                json_data = json.loads(request.body)
                title = json_data.get('title')
                start_date = json_data.get('start_date')
                description = json_data.get('description')
                # Parse time and convert to 'HH:MM:SS' format
                time_str = json_data.get('time')
                time = datetime.strptime(time_str, '%I:%M %p').strftime('%H:%M:%S')
                duration = json_data.get('duration')
                marks = json_data.get('marks')
                rules = json_data.get('rules')
                creator = json_data.get('creator')
                select_students = json_data.get('select_students', [])  # If not provided, default to empty list
                excel_file = None  # Initialize excel_file to None
            else:
                # Handle FormData
                title = request.POST.get('title')
                start_date = request.POST.get('start_date')
                description = request.POST.get('description')
                # Parse time and convert to 'HH:MM:SS' format
                time_str = request.POST.get('time')
                time = datetime.strptime(time_str, '%I:%M %p').strftime('%H:%M:%S')
                duration = request.POST.get('duration')
                marks = request.POST.get('marks')
                rules = request.POST.get('rules')
                creator = request.POST.get('creator')
                select_students = request.POST.get('select_students')
                select_students = [int(user_id) for user_id in select_students.split(',')] if select_students else []

                # Fetching uploaded file
                excel_file = request.FILES.get('file')

            # Input validation
            if not all([title, start_date, time, duration, marks, creator]):
                return JsonResponse({"error": "Missing required fields"}, status=400)

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT user_id FROM user WHERE full_name = %s AND role = 2
                """, [creator])

                creator_data = cursor.fetchone()
                if creator_data:
                    creator_user_id = creator_data[0]
                else:
                    return JsonResponse(
                        {"error": f"No user found for creator: {creator}"}, status=404
                    )

                cursor.execute(
                    """
                    INSERT INTO exams (title, start_date, description, time, duration, marks, rules, creator)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                    [
                        title,
                        start_date,
                        description,
                        time,
                        duration,
                        marks,
                        rules,
                        creator_user_id,
                    ],
                )

                connection.commit()

                exam_id = cursor.lastrowid  # Use last inserted row id

                # Check if any students are selected
                if select_students:
                    for select_student in select_students:
                        cursor.execute("""
                            INSERT INTO candidate_exam (exam_id, user_id, approval_status)
                            VALUES (%s, %s, 'Y')
                        """, [exam_id, select_student])
                        logger.info("Candidate added for exam_id %s and user_id %s", exam_id, select_student)

                if excel_file:
                    # Check if the uploaded file is in the expected Excel format
                    if not excel_file.name.endswith(".xlsx"):
                        return JsonResponse(
                            {
                                "error": "Invalid file format. Please upload an Excel file (.xlsx)"
                            },
                            status=400,
                        )

                    # Attempt to load the workbook
                    try:
                        wb = load_workbook(excel_file)
                    except Exception as e:
                        logger.error("Failed to load Excel file: %s", str(e))
                        return JsonResponse(
                            {"error": f"Failed to load Excel file: {e}"}, status=400
                        )

                    ws = wb.active

                    # Continue processing the workbook data
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        full_name, email = row[:2]
                        cursor.execute("""
                            SELECT user_id FROM user WHERE email = %s
                        """, [email])
                        user_data = cursor.fetchone()
                        if user_data:
                            user_id = user_data[0]
                            cursor.execute("""
                                INSERT INTO candidate_exam (exam_id, user_id, approval_status)
                                VALUES (%s, %s, 'Y')
                            """, [exam_id, user_id])
                            logger.info("Candidate added for exam_id %s and user_id %s", exam_id, user_id)
                        else:
                            logger.warning("User not found for email: %s", email)

            # Email sending block
            student_ids = select_students
            if student_ids:
                try:
                    student_ids = [int(id) for id in student_ids]
                except (ValueError, TypeError):
                    return JsonResponse(
                        {"error": "Student IDs must be integers"}, status=400
                    )

                # Fetch email addresses from the database
                with connection.cursor() as cursor:
                    placeholders = ", ".join(["%s"] * len(student_ids))
                    query = f"SELECT user_id, email FROM user WHERE user_id IN ({placeholders})"
                    cursor.execute(query, student_ids)
                    students = cursor.fetchall()

                if not students:
                    return JsonResponse(
                        {"error": "No matching students found"}, status=404
                    )

                subject = f"Invitation: {title}"
                message = f"""
                            Dear Student,

                            You are invited to participate in the following exam:

                            Title: {title}
                            Date: {start_date}
                            Time: {time}
                            Duration: {duration} minutes

                            Description:
                            {description}

                            Please log in to your account for more details and to confirm your participation:
                            https://aaas.tdtl.world/login/

                            Best regards,
                            From AAAS
                                            """
                from_email = "pr@thedatatechlabs.com"
                successful_emails = []
                failed_emails = []

                for student_id, email in students:
                    to_email = [email]
                    try:
                        send_mail(
                            subject, message, from_email, to_email, fail_silently=False
                        )
                        successful_emails.append(student_id)
                        logger.info(
                            "Email sent successfully to student %s at %s",
                            student_id,
                            email,
                        )
                    except Exception as e:
                        failed_emails.append(student_id)
                        logger.error(
                            "Error sending email to student %s at %s: %s",
                            student_id,
                            email,
                            e,
                        )

                response_data_email = {
                    "message": "Emails processed",
                    "students_notified": successful_emails,
                    "failed_emails": failed_emails,
                }

            response_data = {"status": "Exam added successfully"}

            if response_data_email:
                response_data["email_info"] = response_data_email

            return JsonResponse(response_data, status=200)

        except IntegrityError as e:
            logger.error("Integrity Error: %s", str(e))
            return JsonResponse({"error": f"Integrity Error: {e}"}, status=400)
        except IOError as e:
            logger.error("I/O Error: %s", str(e))
            return JsonResponse({"error": f"I/O Error: {e}"}, status=500)
        except Exception as e:
            logger.error("An error occurred: %s", str(e))
            logger.exception("Detailed exception:")
            return JsonResponse({"error": f"An error occurred: {e}"}, status=500)
